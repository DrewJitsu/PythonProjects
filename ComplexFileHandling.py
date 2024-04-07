'''
Andrew Harris
Complex File Handling
'''

import csv
from os import chdir
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

MYPATH ="C:\\Users\\andhar5173\\Desktop\\CIS123\\Week3"

chdir(MYPATH)

wb=load_workbook("example.xlsx") #loading example workbook
ws = wb.active #reference to active worksheet
rows = ws.max_row
cols = ws.max_column

alphas = []
for charCode in range(65,91): #setting character range to ASCII capital letters
    alphas.append(chr(charCode))

for row in range (1, rows+1): #for each row in the range of rows
    for column in range(cols): #iterate through each col in range of cols
        print(ws[alphas[column]+ str(row)].value, end='\t') #print cell value
    print('\n') #print new line before new cell value


myDataLabels=Reference(ws,min_col=1,min_row=2,max_row=rows) #getting data labels
myData=Reference(ws,min_col=2,min_row=2,max_row=rows) #getting data
xAxisName=ws['A1'].value #getting value from ws @ !A1
yAxisName=ws['B1'].value #getting value from ws @ !B1

myChart=BarChart() #setting type of chart to bar chart

myChart.Title="Fruit Quantities" #setting chart title
myChart.legend=None #removing legend

myChart.add_data(myData) #adding data to be plotted
myChart.set_categories(myDataLabels) #setting values that will appear under bars
myChart.x_axis.title=xAxisName #setting title of x axis
myChart.y_axis.title=yAxisName #setting title of y axis

ws.add_chart(myChart,"A11") #adding chart to sheet one starting at A11 location

wb.save("exampleChart.xlsx") #saving workbook as new excel file
